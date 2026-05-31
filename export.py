"""
export.py — Finance PRO
=======================
Exportação profissional de relatórios financeiros.
- Excel: planilha surreal com capa, KPIs, formatação condicional e tabelas formatadas
- PDF:   relatório mensal com fpdf2
- CSV:   fallback simples
"""

import io
from datetime import date


def gerar_excel(lancamentos: list, investimentos: list, metas: list) -> bytes:
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
            GradientFill
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint

        wb = Workbook()

        # ── Paleta de cores ────────────────────────────────────────────────────
        ROXO_ESC  = "3B0764"
        ROXO      = "7C3AED"
        ROXO_MED  = "A78BFA"
        ROXO_CLR  = "EDE9FE"
        ROXO_ULCL = "F5F3FF"
        VERDE     = "16A34A"
        VERDE_CLR = "DCFCE7"
        VERM      = "DC2626"
        VERM_CLR  = "FEE2E2"
        AMBAR     = "D97706"
        AMBAR_CLR = "FEF3C7"
        CINZA_ESC = "1E1B4B"
        CINZA_MED = "64748B"
        CINZA_CLR = "F8FAFC"
        BRANCO    = "FFFFFF"
        PRETO     = "0F0F0F"

        # ── Bordas ─────────────────────────────────────────────────────────────
        def borda_fina(cor="D1D5DB"):
            s = Side(style="thin", color=cor)
            return Border(left=s, right=s, top=s, bottom=s)

        def borda_media(cor="7C3AED"):
            s = Side(style="medium", color=cor)
            return Border(left=s, right=s, top=s, bottom=s)

        def borda_baixo(cor="7C3AED"):
            return Border(bottom=Side(style="medium", color=cor))

        # ── Helpers de estilo ──────────────────────────────────────────────────
        def fill(hex_cor):
            return PatternFill("solid", fgColor=hex_cor)

        def fonte(bold=False, size=10, cor=PRETO, italic=False, nome="Calibri"):
            return Font(bold=bold, size=size, color=cor, italic=italic, name=nome)

        def alinhar(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def moeda_br(ws, row, col):
            ws.cell(row=row, column=col).number_format = '#.##0,00'

        # ══════════════════════════════════════════════════════════════════════
        # ABA 1 — CAPA / RESUMO EXECUTIVO
        # ══════════════════════════════════════════════════════════════════════
        ws_capa = wb.active
        ws_capa.title = "📊 Resumo Executivo"
        ws_capa.sheet_view.showGridLines = False
        ws_capa.column_dimensions["A"].width = 3
        ws_capa.column_dimensions["B"].width = 28
        ws_capa.column_dimensions["C"].width = 22
        ws_capa.column_dimensions["D"].width = 22
        ws_capa.column_dimensions["E"].width = 22
        ws_capa.column_dimensions["F"].width = 3

        # Fundo geral da capa
        for row in ws_capa.iter_rows(min_row=1, max_row=60, min_col=1, max_col=6):
            for cell in row:
                cell.fill = fill(CINZA_CLR)

        # Banner topo — fundo escuro
        for row in range(1, 9):
            ws_capa.row_dimensions[row].height = 18
            for col in range(1, 7):
                ws_capa.cell(row=row, column=col).fill = fill(CINZA_ESC)

        # Título principal
        ws_capa.merge_cells("B2:E3")
        c = ws_capa["B2"]
        c.value = "FINANCE PRO"
        c.font = Font(bold=True, size=28, color=ROXO_MED, name="Calibri")
        c.alignment = alinhar("left", "center")

        ws_capa.merge_cells("B4:E4")
        c = ws_capa["B4"]
        c.value = "Relatório Financeiro Completo"
        c.font = fonte(size=12, cor=BRANCO, italic=True)
        c.alignment = alinhar("left", "center")

        ws_capa.merge_cells("B5:E5")
        c = ws_capa["B5"]
        c.value = f"Gerado em {date.today().strftime('%d/%m/%Y')}"
        c.font = fonte(size=10, cor="9CA3AF")
        c.alignment = alinhar("left", "center")

        # Linha separadora roxa
        for col in range(1, 7):
            ws_capa.cell(row=8, column=col).fill = fill(ROXO)

        ws_capa.row_dimensions[8].height = 4

        # ── KPIs calculados ────────────────────────────────────────────────────
        total_ent  = sum(t.get("valor", 0) for t in lancamentos if t.get("tipo") == "entrada")
        total_sai  = sum(t.get("valor", 0) for t in lancamentos if t.get("tipo") == "saida")
        saldo      = total_ent - total_sai
        total_inv  = sum(i.get("valor", 0) for i in investimentos)
        patrimonio = saldo + total_inv
        poupanca   = round(saldo / total_ent * 100, 1) if total_ent > 0 else 0

        def fmt_br(v):
            return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        # Linha de rótulos KPI
        ws_capa.row_dimensions[10].height = 14
        ws_capa.row_dimensions[11].height = 36
        ws_capa.row_dimensions[12].height = 20
        ws_capa.row_dimensions[13].height = 10

        kpis = [
            ("RECEITA TOTAL",    fmt_br(total_ent),  VERDE,  VERDE_CLR,  "C"),
            ("DESPESAS TOTAIS",  fmt_br(total_sai),  VERM,   VERM_CLR,   "D"),
            ("SALDO LÍQUIDO",    fmt_br(saldo),      ROXO,   ROXO_CLR,   "E"),
        ]

        for label, valor, cor_val, cor_bg, col in kpis:
            # Caixa KPI
            ws_capa[f"{col}10"].value = label
            ws_capa[f"{col}10"].font = fonte(bold=True, size=8, cor=CINZA_MED)
            ws_capa[f"{col}10"].alignment = alinhar("center")

            ws_capa[f"{col}11"].value = valor
            ws_capa[f"{col}11"].font = Font(bold=True, size=14, color=cor_val, name="Calibri")
            ws_capa[f"{col}11"].alignment = alinhar("center", "center")
            ws_capa[f"{col}11"].fill = fill(cor_bg)
            ws_capa[f"{col}11"].border = borda_media(cor_val)

        # KPIs linha 2
        ws_capa.row_dimensions[15].height = 14
        ws_capa.row_dimensions[16].height = 36
        kpis2 = [
            ("INVESTIMENTOS",    fmt_br(total_inv),   AMBAR,  AMBAR_CLR,  "C"),
            ("PATRIMÔNIO",       fmt_br(patrimonio),  ROXO,   ROXO_CLR,   "D"),
            ("TAXA DE POUPANÇA", f"{poupanca}%",      VERDE,  VERDE_CLR,  "E"),
        ]
        for label, valor, cor_val, cor_bg, col in kpis2:
            ws_capa[f"{col}15"].value = label
            ws_capa[f"{col}15"].font = fonte(bold=True, size=8, cor=CINZA_MED)
            ws_capa[f"{col}15"].alignment = alinhar("center")
            ws_capa[f"{col}16"].value = valor
            ws_capa[f"{col}16"].font = Font(bold=True, size=14, color=cor_val, name="Calibri")
            ws_capa[f"{col}16"].alignment = alinhar("center", "center")
            ws_capa[f"{col}16"].fill = fill(cor_bg)
            ws_capa[f"{col}16"].border = borda_media(cor_val)

        # ── Tabela de categorias ───────────────────────────────────────────────
        cats = {}
        for t in lancamentos:
            if t.get("tipo") == "saida":
                cats[t.get("categoria","Outros")] = cats.get(t.get("categoria","Outros"),0) + t.get("valor",0)
        cats_ord = sorted(cats.items(), key=lambda x: -x[1])

        ws_capa.row_dimensions[19].height = 18
        ws_capa.merge_cells("B19:E19")
        c = ws_capa["B19"]
        c.value = "DESPESAS POR CATEGORIA"
        c.font = Font(bold=True, size=11, color=BRANCO, name="Calibri")
        c.alignment = alinhar("left", "center")
        c.fill = fill(ROXO)

        headers_cat = ["Categoria", "Valor (R$)", "% do Total"]
        for i, h in enumerate(headers_cat, 2):
            col = get_column_letter(i)
            ws_capa[f"{col}20"].value = h
            ws_capa[f"{col}20"].font = fonte(bold=True, size=9, cor=ROXO_ESC)
            ws_capa[f"{col}20"].fill = fill(ROXO_ULCL)
            ws_capa[f"{col}20"].alignment = alinhar("center")
            ws_capa[f"{col}20"].border = borda_fina(ROXO_MED)

        for row_i, (cat, val) in enumerate(cats_ord, 21):
            pct_cat = round(val / total_sai * 100, 1) if total_sai > 0 else 0
            bg = BRANCO if row_i % 2 == 0 else ROXO_ULCL
            dados = [cat, val, f"{pct_cat}%"]
            for col_i, dado in enumerate(dados, 2):
                cell = ws_capa.cell(row=row_i, column=col_i)
                cell.value = dado
                cell.fill = fill(bg)
                cell.border = borda_fina()
                cell.alignment = alinhar("center" if col_i > 2 else "left")
                if col_i == 3:
                    cell.font = Font(bold=True, size=10, color=VERM if pct_cat > 30 else CINZA_MED, name="Calibri")
                else:
                    cell.font = fonte(size=10)
                if col_i == 3 and isinstance(dado, (int, float)):
                    cell.number_format = '#.##0,00'

        # ── Rodapé da capa ─────────────────────────────────────────────────────
        rodape_row = 22 + len(cats_ord) + 1
        ws_capa.row_dimensions[rodape_row].height = 4
        for col in range(1, 7):
            ws_capa.cell(row=rodape_row, column=col).fill = fill(ROXO)

        ws_capa.merge_cells(f"B{rodape_row+1}:E{rodape_row+1}")
        c = ws_capa.cell(row=rodape_row+1, column=2)
        c.value = "Finance PRO — Relatório gerado automaticamente"
        c.font = fonte(size=9, cor=CINZA_MED, italic=True)
        c.alignment = alinhar("center")

        # ══════════════════════════════════════════════════════════════════════
        # ABA 2 — LANÇAMENTOS
        # ══════════════════════════════════════════════════════════════════════
        ws_lanc = wb.create_sheet("💸 Lançamentos")
        ws_lanc.sheet_view.showGridLines = False

        # Larguras
        larguras = {"A":6,"B":14,"C":32,"D":18,"E":14,"F":16,"G":8}
        for col, w in larguras.items():
            ws_lanc.column_dimensions[col].width = w

        # Header banner
        ws_lanc.row_dimensions[1].height = 4
        for col in range(1, 8):
            ws_lanc.cell(row=1, column=col).fill = fill(CINZA_ESC)

        ws_lanc.merge_cells("A2:G2")
        ws_lanc.row_dimensions[2].height = 28
        c = ws_lanc["A2"]
        c.value = "💸  LANÇAMENTOS FINANCEIROS"
        c.font = Font(bold=True, size=14, color=BRANCO, name="Calibri")
        c.alignment = alinhar("left", "center")
        c.fill = fill(CINZA_ESC)

        ws_lanc.merge_cells("A3:G3")
        ws_lanc.row_dimensions[3].height = 16
        c = ws_lanc["A3"]
        c.value = f"Total: {len(lancamentos)} registros  |  Receitas: {fmt_br(total_ent)}  |  Despesas: {fmt_br(total_sai)}  |  Saldo: {fmt_br(saldo)}"
        c.font = fonte(size=9, cor="9CA3AF", italic=True)
        c.alignment = alinhar("left", "center")
        c.fill = fill(CINZA_ESC)

        ws_lanc.row_dimensions[4].height = 4
        for col in range(1, 8):
            ws_lanc.cell(row=4, column=col).fill = fill(ROXO)

        # Cabeçalho da tabela
        headers = ["#", "Data", "Descrição", "Categoria", "Tipo", "Valor (R$)", "Ícone"]
        ws_lanc.row_dimensions[5].height = 22
        for i, h in enumerate(headers, 1):
            cell = ws_lanc.cell(row=5, column=i)
            cell.value = h
            cell.font = Font(bold=True, size=10, color=BRANCO, name="Calibri")
            cell.fill = fill(ROXO)
            cell.alignment = alinhar("center", "center")
            cell.border = Border(
                bottom=Side(style="medium", color=ROXO_MED),
                right=Side(style="thin", color="5B21B6")
            )

        # Dados
        for idx, t in enumerate(lancamentos, 1):
            row = idx + 5
            ws_lanc.row_dimensions[row].height = 17
            is_ent = t.get("tipo") == "entrada"
            bg = VERDE_CLR if is_ent else VERM_CLR if idx % 2 == 0 else (DCFCE7 := "F0FDF4") if is_ent else "FFF1F2"
            bg = ("F0FDF4" if is_ent else "FFF1F2") if idx % 2 == 1 else ("DCFCE7" if is_ent else "FEE2E2")

            dados_row = [
                idx,
                str(t.get("data",""))[:10],
                t.get("nome",""),
                t.get("categoria",""),
                "✅ Entrada" if is_ent else "❌ Saída",
                t.get("valor", 0),
                t.get("icone",""),
            ]
            for col_i, val in enumerate(dados_row, 1):
                cell = ws_lanc.cell(row=row, column=col_i)
                cell.value = val
                cell.fill = fill(bg)
                cell.border = borda_fina("E5E7EB")
                cell.font = fonte(size=9, cor=VERDE if is_ent and col_i==6 else VERM if not is_ent and col_i==6 else PRETO,
                                  bold=(col_i==6))
                cell.alignment = alinhar("center" if col_i in [1,2,5,6,7] else "left")
                if col_i == 6:
                    cell.number_format = '#.##0,00'

        # Linha de totais
        total_row = len(lancamentos) + 6
        ws_lanc.row_dimensions[total_row].height = 20
        ws_lanc.merge_cells(f"A{total_row}:E{total_row}")
        c = ws_lanc.cell(row=total_row, column=1)
        c.value = f"TOTAL — {len(lancamentos)} lançamentos"
        c.font = Font(bold=True, size=10, color=BRANCO, name="Calibri")
        c.fill = fill(ROXO_ESC)
        c.alignment = alinhar("right", "center")
        c.border = borda_media()

        c_tot = ws_lanc.cell(row=total_row, column=6)
        c_tot.value = total_ent - total_sai
        c_tot.font = Font(bold=True, size=11, color=VERDE if saldo >= 0 else VERM, name="Calibri")
        c_tot.fill = fill(ROXO_ESC)
        c_tot.alignment = alinhar("center", "center")
        c_tot.number_format = '#.##0,00'
        c_tot.border = borda_media()

        # ══════════════════════════════════════════════════════════════════════
        # ABA 3 — INVESTIMENTOS
        # ══════════════════════════════════════════════════════════════════════
        ws_inv = wb.create_sheet("📈 Investimentos")
        ws_inv.sheet_view.showGridLines = False

        for col, w in {"A":6,"B":32,"C":18,"D":16,"E":14}.items():
            ws_inv.column_dimensions[col].width = w

        # Banner
        for col in range(1, 6):
            ws_inv.cell(row=1, column=col).fill = fill(CINZA_ESC)
        ws_inv.row_dimensions[1].height = 4

        ws_inv.merge_cells("A2:E2")
        ws_inv.row_dimensions[2].height = 28
        c = ws_inv["A2"]
        c.value = "📈  CARTEIRA DE INVESTIMENTOS"
        c.font = Font(bold=True, size=14, color=BRANCO, name="Calibri")
        c.alignment = alinhar("left", "center")
        c.fill = fill(CINZA_ESC)

        ws_inv.merge_cells("A3:E3")
        ws_inv.row_dimensions[3].height = 16
        c = ws_inv["A3"]
        c.value = f"Total investido: {fmt_br(total_inv)}  |  {len(investimentos)} ativos"
        c.font = fonte(size=9, cor="9CA3AF", italic=True)
        c.alignment = alinhar("left")
        c.fill = fill(CINZA_ESC)

        ws_inv.row_dimensions[4].height = 4
        for col in range(1, 6):
            ws_inv.cell(row=4, column=col).fill = fill(AMBAR)

        headers_inv = ["#", "Ativo", "Valor (R$)", "Variação (%)", "% do Portfolio"]
        ws_inv.row_dimensions[5].height = 22
        for i, h in enumerate(headers_inv, 1):
            cell = ws_inv.cell(row=5, column=i)
            cell.value = h
            cell.font = Font(bold=True, size=10, color=BRANCO, name="Calibri")
            cell.fill = fill(AMBAR)
            cell.alignment = alinhar("center")
            cell.border = Border(bottom=Side(style="medium", color="FCD34D"))

        for idx, inv in enumerate(investimentos, 1):
            row = idx + 5
            ws_inv.row_dimensions[row].height = 18
            pct_port = round(inv.get("valor",0) / total_inv * 100, 1) if total_inv > 0 else 0
            try:
                var = float(str(inv.get("variacao",0)).replace("%","").replace("+","").replace(",","."))
            except:
                var = 0.0
            bg = AMBAR_CLR if idx % 2 == 0 else BRANCO
            dados_inv = [idx, inv.get("nome",""), inv.get("valor",0), var, pct_port]
            for col_i, val in enumerate(dados_inv, 1):
                cell = ws_inv.cell(row=row, column=col_i)
                cell.value = val
                cell.fill = fill(bg)
                cell.border = borda_fina("E5E7EB")
                cell.alignment = alinhar("left" if col_i==2 else "center")
                if col_i == 3:
                    cell.number_format = '#.##0,00'
                    cell.font = Font(bold=True, size=10, color=AMBAR, name="Calibri")
                elif col_i == 4:
                    cell.number_format = '+0.00;-0.00;0.00'
                    cell.font = Font(bold=True, size=10,
                                     color=VERDE if var >= 0 else VERM, name="Calibri")
                elif col_i == 5:
                    cell.value = f"{pct_port}%"
                    cell.font = fonte(size=10)
                else:
                    cell.font = fonte(size=10)

        # Total investimentos
        tot_row_inv = len(investimentos) + 6
        ws_inv.row_dimensions[tot_row_inv].height = 20
        ws_inv.merge_cells(f"A{tot_row_inv}:B{tot_row_inv}")
        c = ws_inv.cell(row=tot_row_inv, column=1)
        c.value = "TOTAL DO PORTFÓLIO"
        c.font = Font(bold=True, size=10, color=BRANCO, name="Calibri")
        c.fill = fill(CINZA_ESC)
        c.alignment = alinhar("right", "center")

        c2 = ws_inv.cell(row=tot_row_inv, column=3)
        c2.value = total_inv
        c2.font = Font(bold=True, size=12, color=AMBAR, name="Calibri")
        c2.fill = fill(CINZA_ESC)
        c2.number_format = '#.##0,00'
        c2.alignment = alinhar("center", "center")

        # ══════════════════════════════════════════════════════════════════════
        # ABA 4 — METAS
        # ══════════════════════════════════════════════════════════════════════
        ws_metas = wb.create_sheet("🎯 Metas")
        ws_metas.sheet_view.showGridLines = False

        for col, w in {"A":6,"B":30,"C":16,"D":16,"E":14,"F":16}.items():
            ws_metas.column_dimensions[col].width = w

        for col in range(1, 7):
            ws_metas.cell(row=1, column=col).fill = fill(CINZA_ESC)
        ws_metas.row_dimensions[1].height = 4

        ws_metas.merge_cells("A2:F2")
        ws_metas.row_dimensions[2].height = 28
        c = ws_metas["A2"]
        c.value = "🎯  METAS FINANCEIRAS"
        c.font = Font(bold=True, size=14, color=BRANCO, name="Calibri")
        c.alignment = alinhar("left", "center")
        c.fill = fill(CINZA_ESC)

        ws_metas.row_dimensions[3].height = 4
        for col in range(1, 7):
            ws_metas.cell(row=3, column=col).fill = fill(VERDE)

        headers_m = ["#", "Meta", "Atual (R$)", "Objetivo (R$)", "Progresso %", "Status"]
        ws_metas.row_dimensions[4].height = 22
        for i, h in enumerate(headers_m, 1):
            cell = ws_metas.cell(row=4, column=i)
            cell.value = h
            cell.font = Font(bold=True, size=10, color=BRANCO, name="Calibri")
            cell.fill = fill(VERDE)
            cell.alignment = alinhar("center")
            cell.border = Border(bottom=Side(style="medium", color="4ADE80"))

        for idx, m in enumerate(metas, 1):
            row = idx + 4
            ws_metas.row_dimensions[row].height = 20
            pct = round(m.get("atual",0) / m.get("total",1) * 100, 1) if m.get("total",1) > 0 else 0
            pct = min(pct, 100)
            status = "✅ Concluída" if pct >= 100 else ("🔥 Quase lá!" if pct >= 75 else ("📈 Em andamento" if pct >= 25 else "🚀 Iniciando"))
            cor_status = VERDE if pct >= 100 else (AMBAR if pct >= 75 else ROXO)
            bg = VERDE_CLR if pct >= 100 else (AMBAR_CLR if idx % 2 == 0 else BRANCO)

            dados_m = [idx, m.get("nome",""), m.get("atual",0), m.get("total",0), pct, status]
            for col_i, val in enumerate(dados_m, 1):
                cell = ws_metas.cell(row=row, column=col_i)
                cell.value = val
                cell.fill = fill(bg)
                cell.border = borda_fina("E5E7EB")
                cell.alignment = alinhar("left" if col_i==2 else "center")
                if col_i in [3, 4]:
                    cell.number_format = '#.##0,00'
                    cell.font = fonte(size=10, bold=True)
                elif col_i == 5:
                    cell.number_format = '0.0"%"'
                    cell.font = Font(bold=True, size=11,
                                     color=VERDE if pct>=100 else AMBAR if pct>=50 else VERM,
                                     name="Calibri")
                elif col_i == 6:
                    cell.font = Font(bold=True, size=9, color=cor_status, name="Calibri")
                else:
                    cell.font = fonte(size=10)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except ImportError as e:
        raise ImportError(f"Instale openpyxl: pip install openpyxl. Detalhe: {e}")


def gerar_pdf(
    lancamentos: list,
    investimentos: list,
    metas: list,
    nome_usuario: str = "Usuário",
    mes: int = None,
    ano: int = None,
) -> bytes:
    """
    Gera relatório PDF mensal com KPIs, tabelas e resumo de metas.
    Requer fpdf2: pip install fpdf2
    """
    try:
        from fpdf import FPDF

        mes  = mes  or date.today().month
        ano  = ano  or date.today().year
        MESES = ["Janeiro","Fevereiro","Marco","Abril","Maio","Junho",
                  "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        def s(txt):
            """Converte texto para latin-1 seguro para fpdf2."""
            return str(txt).encode("latin-1", errors="replace").decode("latin-1")
        mes_nome = MESES[mes - 1]

        total_ent = sum(t.get("valor",0) for t in lancamentos if t.get("tipo")=="entrada")
        total_sai = sum(t.get("valor",0) for t in lancamentos if t.get("tipo")=="saida")
        saldo     = total_ent - total_sai
        total_inv = sum(i.get("valor",0) for i in investimentos)
        poupanca  = round(saldo / total_ent * 100, 1) if total_ent > 0 else 0

        def fmtv(v):
            return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        pdf.set_fill_color(30, 27, 75)
        pdf.rect(0, 0, 210, 40, "F")

        pdf.set_xy(10, 8)
        pdf.set_font("Helvetica", "B", 22)
        pdf.set_text_color(167, 139, 250)
        pdf.cell(0, 10, "FINANCE PRO", ln=False)

        pdf.set_xy(10, 20)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(200, 200, 220)
        pdf.cell(0, 8, s(f"Relatorio de {mes_nome}/{ano}  -  {nome_usuario}"), ln=True)

        pdf.set_xy(10, 30)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(150, 150, 180)
        pdf.cell(0, 6, s(f"Gerado em {date.today().strftime('%d/%m/%Y')}"), ln=True)

        # Linha roxa
        pdf.set_fill_color(124, 58, 237)
        pdf.rect(0, 40, 210, 2, "F")

        pdf.ln(10)

        # ── KPIs ──────────────────────────────────────────────────────────────
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(30, 27, 75)
        pdf.cell(0, 6, s("RESUMO DO PERIODO"), ln=True)
        pdf.ln(2)

        kpis_pdf = [
            (s("RECEITA"),    s(fmtv(total_ent)), (22, 163, 74)),
            (s("DESPESAS"),   s(fmtv(total_sai)), (220, 38, 38)),
            (s("SALDO"),      s(fmtv(saldo)),     (124, 58, 237)),
            (s("POUPANCA"),   s(f"{poupanca}%"),  (8, 145, 178)),
        ]

        box_w = 44
        for i, (label, valor, cor) in enumerate(kpis_pdf):
            x = 10 + i * (box_w + 3)
            y = pdf.get_y()
            pdf.set_fill_color(248, 250, 252)
            pdf.rect(x, y, box_w, 18, "F")
            pdf.set_fill_color(*cor)
            pdf.rect(x, y, 3, 18, "F")
            pdf.set_xy(x + 5, y + 2)
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(box_w - 6, 4, s(label), ln=True)
            pdf.set_xy(x + 5, y + 8)
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(*cor)
            pdf.cell(box_w - 6, 6, s(valor), ln=False)

        pdf.ln(25)

        # ── Tabela de lançamentos ──────────────────────────────────────────────
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(30, 27, 75)
        pdf.cell(0, 6, s(f"LANCAMENTOS ({len(lancamentos)} registros)"), ln=True)
        pdf.ln(1)

        # Cabeçalho da tabela
        col_ws = [18, 65, 35, 25, 30]
        headers_pdf = [s("Data"), s("Descricao"), s("Categoria"), s("Tipo"), s("Valor (R$)")]
        pdf.set_fill_color(124, 58, 237)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for i, (h, w) in enumerate(zip(headers_pdf, col_ws)):
            pdf.cell(w, 7, s(h), border=0, fill=True, align="C")
        pdf.ln()

        # Linhas de dados
        for idx, t in enumerate(lancamentos):
            is_ent = t.get("tipo") == "entrada"
            if idx % 2 == 0:
                pdf.set_fill_color(245, 243, 255)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(15, 15, 15)
            pdf.set_font("Helvetica", "", 8)

            pdf.cell(col_ws[0], 6, s(str(t.get("data",""))[:10]), border=0, fill=True, align="C")
            pdf.cell(col_ws[1], 6, s(t.get("nome","")[:35]), border=0, fill=True)
            pdf.cell(col_ws[2], 6, s(t.get("categoria","")), border=0, fill=True, align="C")

            tipo_label = s("Entrada") if is_ent else s("Saida")
            pdf.set_text_color(22, 163, 74) if is_ent else pdf.set_text_color(220, 38, 38)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(col_ws[3], 6, tipo_label, border=0, fill=True, align="C")

            pdf.set_text_color(22, 163, 74) if is_ent else pdf.set_text_color(220, 38, 38)
            pdf.cell(col_ws[4], 6, s(fmtv(t.get("valor",0))), border=0, fill=True, align="R")
            pdf.ln()

        # Total
        pdf.set_fill_color(30, 27, 75)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(sum(col_ws[:4]), 7, s("SALDO DO PERIODO"), fill=True, align="R")
        pdf.set_text_color(74, 222, 128) if saldo >= 0 else pdf.set_text_color(248, 113, 113)
        pdf.cell(col_ws[4], 7, s(fmtv(saldo)), fill=True, align="R")
        pdf.ln(10)

        # ── Metas ─────────────────────────────────────────────────────────────
        if metas:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(30, 27, 75)
            pdf.cell(0, 6, s("METAS FINANCEIRAS"), ln=True)
            pdf.ln(1)

            for m in metas:
                pct = min(round(m.get("atual",0) / m.get("total",1) * 100), 100)
                pdf.set_font("Helvetica", "B", 9)
                pdf.set_text_color(15, 15, 15)
                pdf.cell(100, 5, s(m.get("nome","")), ln=False)
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(100, 116, 139)
                pdf.cell(0, 5, s(f"{fmtv(m.get('atual',0))} / {fmtv(m.get('total',0))}  ({pct}%)"), ln=True)

                # Barra de progresso
                bar_x = pdf.get_x() + 10
                bar_y = pdf.get_y()
                bar_w = 170
                pdf.set_fill_color(229, 231, 235)
                pdf.rect(bar_x, bar_y, bar_w, 3, "F")
                pdf.set_fill_color(124, 58, 237) if pct < 100 else pdf.set_fill_color(22, 163, 74)
                pdf.rect(bar_x, bar_y, bar_w * pct / 100, 3, "F")
                pdf.ln(6)

        # ── Rodapé ────────────────────────────────────────────────────────────
        pdf.set_y(-20)
        pdf.set_fill_color(124, 58, 237)
        pdf.rect(0, pdf.get_y(), 210, 0.5, "F")
        pdf.ln(2)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(150, 150, 180)
        pdf.cell(0, 5, s("Finance PRO - Relatorio gerado automaticamente"), align="C")

        return bytes(pdf.output())

    except ImportError:
        raise ImportError("Instale fpdf2: pip install fpdf2")


def gerar_csv(lancamentos: list) -> str:
    linhas = ["Data,Descrição,Categoria,Tipo,Valor"]
    for t in lancamentos:
        linhas.append(
            f"{str(t.get('data',''))[:10]},"
            f"{t.get('nome','')},"
            f"{t.get('categoria','')},"
            f"{'Entrada' if t.get('tipo')=='entrada' else 'Saída'},"
            f"{t.get('valor',0)}"
        )
    return "\n".join(linhas)
